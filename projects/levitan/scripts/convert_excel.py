"""Скрипт конвертации Excel файлов в CSV для кампаний."""

import os
import csv
import logging
from pathlib import Path
import openpyxl

logger = logging.getLogger(__name__)


def convert_excel_to_csv(
    excel_path: Path,
    output_dir: Path,
    sheet_name: str = None
) -> Path:
    """
    Конвертация Excel файла в CSV.
    
    Args:
        excel_path: Путь к Excel файлу
        output_dir: Директория для вывода
        sheet_name: Имя листа (если None - активный)
        
    Returns:
        Путь к CSV файлу
    """
    try:
        # Читаем Excel
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Получаем заголовки
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(str(cell) if cell else f"col_{len(headers)}")
        
        # Читаем данные
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(row)
        
        wb.close()
        
        # Формируем имя CSV
        csv_name = excel_path.stem + ".csv"
        csv_path = output_dir / csv_name
        
        # Записываем CSV
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)
        
        logger.info(f"Converted {excel_path.name} -> {csv_name} ({len(rows)} rows)")
        return csv_path
        
    except Exception as e:
        logger.error(f"Failed to convert {excel_path}: {e}")
        raise


def convert_all_excel_in_dir(
    input_dir: Path,
    output_dir: Path
) -> list[Path]:
    """
    Конвертация всех Excel файлов в директории.
    
    Args:
        input_dir: Директория с Excel файлами
        output_dir: Директория для CSV файлов
        
    Returns:
        Список путей к CSV файлам
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    
    csv_files = []
    
    for excel_file in input_dir.glob("*.xlsx"):
        if not excel_file.name.startswith("~"):
            try:
                csv_path = convert_excel_to_csv(excel_file, output_dir)
                csv_files.append(csv_path)
            except Exception as e:
                logger.error(f"Failed to convert {excel_file}: {e}")
    
    logger.info(f"Converted {len(csv_files)} files")
    return csv_files


def merge_csv_files(
    csv_files: list[Path],
    output_path: Path
) -> Path:
    """
    Объединение нескольких CSV файлов в один.
    
    Args:
        csv_files: Список путей к CSV файлам
        output_path: Путь к выходному файлу
        
    Returns:
        Путь к объединенному файлу
    """
    all_rows = []
    headers = None
    
    for csv_file in csv_files:
        try:
            with open(csv_file, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                
                if headers is None:
                    headers = reader.fieldnames
                
                for row in reader:
                    all_rows.append(row)
            
            logger.debug(f"Read {csv_file.name}: {len(all_rows)} total rows")
            
        except Exception as e:
            logger.error(f"Failed to read {csv_file}: {e}")
    
    if headers is None:
        headers = []
    
    # Записываем объединенный файл
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(all_rows)
    
    logger.info(f"Merged {len(csv_files)} files -> {output_path} ({len(all_rows)} rows)")
    return output_path


if __name__ == "__main__":
    # Настройка логирования
    logging.basicConfig(level=logging.INFO)
    
    # Пути
    base_dir = Path(__file__).resolve().parent.parent.parent
    campaigns_dir = base_dir / "data" / "campaigns"
    excel_dir = campaigns_dir / "2026 "
    csv_dir = campaigns_dir / "csv"
    
    # Конвертируем все Excel файлы
    csv_files = convert_all_excel_in_dir(excel_dir, csv_dir)
    
    # Объединяем в один файл
    if csv_files:
        merged_path = csv_dir / "all_contacts_2026.csv"
        merge_csv_files(csv_files, merged_path)
        print(f"Готово! Объединенный файл: {merged_path}")
    else:
        print("Excel файлы не найдены")
