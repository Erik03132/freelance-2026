#!/usr/bin/env python3
"""
Создание Excel-калькулятора финансовой модели для ООО ветеранов СВО.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows

def create_financial_model():
    """Создать Excel-файл с финансовой моделью."""
    
    wb = Workbook()
    
    # Стили
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1F4E79")
    number_format = '#,##0.00'
    currency_format = '#,##0.00 "₽"'
    percent_format = '0.0%'
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==================== ЛИСТ 1: Стартовые вложения ====================
    ws_startup = wb.active
    ws_startup.title = "Стартовые вложения"
    
    startup_data = [
        ["Стартовые вложения ООО «Ветеран-Сервис»"],
        [],
        ["Статья расходов", "Сумма (руб.)", "Примечание"],
        ["Регистрация ООО, печать, счёт", 10000, "Госпошлина + нотариус + печать"],
        ["Ремонт и обустройство офиса", 80000, "Мебель, свет, косметика"],
        ["Компьютеры и оргтехника (3 шт.)", 120000, "Ноутбуки + принтер"],
        ["Веб-сайт и интернет-магазин", 60000, "Разработка + домен + хостинг (год)"],
        ["Закупка первой партии мерча", 100000, "Товар для витрины"],
        ["Реклама и продвижение (старт)", 50000, "Контекст, соцсети, визитки"],
        ["Лицензии на ПО (год)", 30000, "Zoom, конструктор курсов, CRM"],
        ["Резервный фонд", 50000, "На непредвиденные расходы"],
        ["ИТОГО", None, ""],
    ]
    
    for row_idx, row_data in enumerate(startup_data, start=1):
        row = ws_startup.append(row_data)
        # Форматирование заголовков
        if row_idx == 3:
            for cell in ws_startup[row_idx]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
        # Форматирование заголовка
        elif row_idx == 1:
            ws_startup['A1'].font = title_font
            ws_startup['A1'].alignment = Alignment(horizontal='left')
        # Форматирование итоговой строки
        elif row_idx == 12:
            ws_startup['A12'].font = Font(bold=True, size=12)
            ws_startup['B12'].font = Font(bold=True, size=12)
            ws_startup['B12'].number_format = currency_format
            ws_startup['B12'].border = thin_border
        # Обычные строки
        elif row_idx > 3 and row_idx < 12:
            for cell in ws_startup[row_idx]:
                cell.border = thin_border
            if row_idx == 4:
                ws_startup['A4'].border = Border(left=thin_border.left, right=thin_border.right, top=thin_border.top, bottom=thin_border.bottom)
    
    # Формула для ИТОГО
    ws_startup['B12'] = "=SUM(B4:B11)"
    ws_startup['B12'].number_format = currency_format
    
    # Ширина колонок
    ws_startup.column_dimensions['A'].width = 45
    ws_startup.column_dimensions['B'].width = 18
    ws_startup.column_dimensions['C'].width = 50
    
    # ==================== ЛИСТ 2: Ежемесячные расходы ====================
    ws_expenses = wb.create_sheet(title="Ежемесячные расходы")
    
    expenses_data = [
        ["Ежемесячные расходы ООО «Ветеран-Сервис»"],
        [],
        ["Статья расходов", "Сумма (руб./мес.)", "Примечание", "Постоянные/Переменные"],
        ["ФОТ (3 человека)", 210000, "70 000 руб./чел.", "Постоянные"],
        ["— Юрист-консультант", 70000, "Оклад + премия", ""],
        ["— Психолог", 70000, "Оклад + премия", ""],
        ["— Менеджер (магазин + курсы)", 70000, "Оклад + премия", ""],
        ["Налоги на ФОТ (УСН 6% + взносы ~30%)", 42000, "Зависит от системы налогообложения", "Постоянные"],
        ["Аренда офиса (30 м², Раменский район)", 25000, "~800–900 руб./м²", "Постоянные"],
        ["Коммунальные услуги", 5000, "Электричество, вода, уборка", "Постоянные"],
        ["Интернет и связь", 3000, "Оптика + мобильная связь", "Постоянные"],
        ["Бухгалтерское сопровождение", 10000, "Аутсорсинг", "Постоянные"],
        ["Хостинг, домен, поддержка сайта", 5000, "Ежемесячно", "Постоянные"],
        ["Реклама и продвижение", 30000, "Контекст, таргет, SEO", "Постоянные"],
        ["Закупка мерча (по факту продаж)", 64000, "Себестоимость товаров", "Переменные"],
        ["Амортизация оборудования", 5000, "3 года", "Постоянные"],
        ["Прочие расходы", 10000, "Канцелярия, вода, кофе", "Постоянные"],
        ["ИТОГО ежемесячно", None, "", ""],
    ]
    
    for row_idx, row_data in enumerate(expenses_data, start=1):
        ws_expenses.append(row_data)
        if row_idx == 3:
            for cell in ws_expenses[row_idx]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
        elif row_idx == 1:
            ws_expenses['A1'].font = title_font
        elif row_idx == 18:
            ws_expenses['A18'].font = Font(bold=True, size=12)
            ws_expenses['B18'].font = Font(bold=True, size=12)
            ws_expenses['B18'].number_format = currency_format
            ws_expenses['B18'].border = thin_border
        elif row_idx > 3 and row_idx < 18:
            for cell in ws_expenses[row_idx]:
                cell.border = thin_border
    
    ws_expenses['B18'] = "=SUM(B4:B17)"
    ws_expenses['B18'].number_format = currency_format
    
    ws_expenses.column_dimensions['A'].width = 50
    ws_expenses.column_dimensions['B'].width = 20
    ws_expenses.column_dimensions['C'].width = 40
    ws_expenses.column_dimensions['D'].width = 18
    
    # ==================== ЛИСТ 3: Доходы по направлениям ====================
    ws_revenue = wb.create_sheet(title="Доходы и услуги")
    
    # Юридические услуги
    ws_revenue['A1'] = "Юридические услуги"
    ws_revenue['A1'].font = title_font
    legal_header = ["Услуга", "Цена (руб.)", "План в месяц (шт.)", "Выручка (руб.)", "Себестоимость (%)", "Валовая прибыль"]
    legal_data = [
        ["Разовая консультация", 3000, 20, None, 17, None],
        ["Письменное заключение", 5000, 10, None, 6, None],
        ["Сопровождение дела", 25000, 3, None, 8, None],
        ["Подготовка документов", 7000, 8, None, 7, None],
        ["Представительство в суде", 15000, 4, None, 7, None],
        ["ИТОГО юридические", None, None, None, None, None],
    ]
    
    ws_revenue.append([])
    ws_revenue.append(legal_header)
    for row in legal_data:
        ws_revenue.append(row)
    
    # Формулы для юридических услуг
    for row_idx in range(4, 9):
        ws_revenue[f'D{row_idx}'] = f"=B{row_idx}*C{row_idx}"
        ws_revenue[f'F{row_idx}'] = f"=D{row_idx}*(1-E{row_idx}/100)"
    
    ws_revenue['D9'] = "=SUM(D4:D8)"
    ws_revenue['F9'] = "=SUM(F4:F8)"
    
    # Психологическая помощь
    ws_revenue['A12'] = "Психологическая помощь"
    ws_revenue['A12'].font = title_font
    psych_header = ["Услуга", "Цена (руб.)", "План в месяц (шт./чел.)", "Выручка (руб.)", "Себестоимость (%)", "Валовая прибыль"]
    psych_data = [
        ["Индивидуальная сессия", 2500, 40, None, 8, None],
        ["Семейная консультация", 4000, 8, None, 8, None],
        ["Групповая терапия", 1500, 12, None, 33, None],
        ["Курс адаптации (4 сессии)", 8000, 5, None, 13, None],
        ["ИТОГО психология", None, None, None, None, None],
    ]
    
    ws_revenue.append([])
    ws_revenue.append(psych_header)
    for row in psych_data:
        ws_revenue.append(row)
    
    for row_idx in range(15, 19):
        ws_revenue[f'D{row_idx}'] = f"=B{row_idx}*C{row_idx}"
        ws_revenue[f'F{row_idx}'] = f"=D{row_idx}*(1-E{row_idx}/100)"
    
    ws_revenue['D19'] = "=SUM(D15:D18)"
    ws_revenue['F19'] = "=SUM(F15:F18)"
    
    # Онлайн-курсы
    ws_revenue['A22'] = "Онлайн-курсы"
    ws_revenue['A22'].font = title_font
    courses_header = ["Курс", "Цена (руб.)", "План в месяц (студентов)", "Выручка (руб.)", "Себестоимость (%)", "Валовая прибыль"]
    courses_data = [
        ["Стрессоустойчивость и лидерство", 12000, 8, None, 17, None],
        ["HR для малого бизнеса", 18000, 5, None, 17, None],
        ["Безопасность и ЧС", 10000, 6, None, 15, None],
        ["Основы предпринимательства", 20000, 4, None, 20, None],
        ["ИТОГО курсы", None, None, None, None, None],
    ]
    
    ws_revenue.append([])
    ws_revenue.append(courses_header)
    for row in courses_data:
        ws_revenue.append(row)
    
    for row_idx in range(25, 29):
        ws_revenue[f'D{row_idx}'] = f"=B{row_idx}*C{row_idx}"
        ws_revenue[f'F{row_idx}'] = f"=D{row_idx}*(1-E{row_idx}/100)"
    
    ws_revenue['D29'] = "=SUM(D25:D28)"
    ws_revenue['F29'] = "=SUM(F25:F28)"
    
    # Магазин мерча
    ws_revenue['A32'] = "Онлайн-магазин мерча"
    ws_revenue['A32'].font = title_font
    merch_header = ["Товар", "Цена (руб.)", "План в месяц (шт.)", "Выручка (руб.)", "Себестоимость (руб./шт.)", "Валовая прибыль"]
    merch_data = [
        ["Шеврон (кастомный)", 450, 100, None, 150, None],
        ["Патч на рюкзак", 300, 80, None, 100, None],
        ["Футболка с принтом", 1500, 40, None, 600, None],
        ["Худи с символикой", 3000, 15, None, 1200, None],
        ["Кепка/бейсболка", 1000, 25, None, 400, None],
        ["Набор (шеврон+патч+футболка)", 2200, 20, None, 850, None],
        ["ИТОГО магазин", None, None, None, None, None],
    ]
    
    ws_revenue.append([])
    ws_revenue.append(merch_header)
    for row in merch_data:
        ws_revenue.append(row)
    
    for row_idx in range(35, 41):
        ws_revenue[f'D{row_idx}'] = f"=B{row_idx}*C{row_idx}"
        ws_revenue[f'F{row_idx}'] = f"=D{row_idx}-(E{row_idx}*C{row_idx})"
    
    ws_revenue['D41'] = "=SUM(D35:D40)"
    ws_revenue['F41'] = "=SUM(F35:F40)"
    
    # ИТОГО общая выручка
    ws_revenue['A44'] = "ОБЩАЯ ВЫРУЧКА"
    ws_revenue['A44'].font = Font(bold=True, size=12)
    ws_revenue['D44'] = "=D9+D19+D29+D41"
    ws_revenue['D44'].font = Font(bold=True, size=12)
    ws_revenue['D44'].number_format = currency_format
    ws_revenue['F44'] = "=F9+F19+F29+F41"
    ws_revenue['F44'].font = Font(bold=True, size=12)
    ws_revenue['F44'].number_format = currency_format
    ws_revenue['E44'] = "Общая валовая прибыль"
    
    # Форматирование
    for row in range(3, 46):
        for col in range(1, 7):
            cell = ws_revenue.cell(row=row, column=col)
            if cell.value is not None:
                cell.border = thin_border
    
    for header_row in [3, 13, 23, 33]:
        for col in range(1, 7):
            cell = ws_revenue.cell(row=header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
    
    ws_revenue.column_dimensions['A'].width = 40
    ws_revenue.column_dimensions['B'].width = 15
    ws_revenue.column_dimensions['C'].width = 18
    ws_revenue.column_dimensions['D'].width = 18
    ws_revenue.column_dimensions['E'].width = 20
    ws_revenue.column_dimensions['F'].width = 18
    
    # ==================== ЛИСТ 4: Точка безубыточности ====================
    ws_breakeven = wb.create_sheet(title="Точка безубыточности")
    
    ws_breakeven['A1'] = "Расчёт точки безубыточности"
    ws_breakeven['A1'].font = title_font
    
    breakeven_header = ["Показатель", "Значение", "Формула/Примечание"]
    breakeven_data = [
        ["Постоянные расходы (в месяц)", None, "Из листа расходов (без закупки мерча)"],
        ["Средняя маржинальность", None, "Валовая прибыль / Выручка"],
        ["Точка безубыточности (выручка)", None, "Постоянные расходы / Маржинальность"],
        [],
        ["Расчёт постоянных расходов:"],
        ["ФОТ + налоги", None, "210 000 + 42 000"],
        ["Аренда + коммунальные", None, "25 000 + 5 000"],
        ["Интернет + связь", None, "3 000"],
        ["Бухгалтерия", None, "10 000"],
        ["Хостинг и ПО", None, "5 000"],
        ["Реклама", None, "30 000"],
        ["Прочие", None, "10 000"],
        ["ИТОГО постоянные", None, ""],
        [],
        ["Расчёт маржинальности:"],
        ["Общая выручка", None, "Из листа доходов"],
        ["Общая валовая прибыль", None, "Из листа доходов"],
        ["Маржинальность %", None, "Валовая прибыль / Выручка * 100"],
    ]
    
    ws_breakeven.append([])
    ws_breakeven.append(breakeven_header)
    for row in breakeven_data:
        ws_breakeven.append(row)
    
    # Формулы
    ws_breakeven['B4'] = "='Ежемесячные расходы'!B18-'Ежемесячные расходы'!B15"  # Постоянные без мерча
    ws_breakeven['B4'].number_format = currency_format
    
    ws_breakeven['B5'] = "='Доходы и услуги'!F44/'Доходы и услуги'!D44"
    ws_breakeven['B5'].number_format = percent_format
    
    ws_breakeven['B6'] = "=B4/B5"
    ws_breakeven['B6'].font = Font(bold=True, size=12)
    ws_breakeven['B6'].number_format = currency_format
    
    ws_breakeven['B13'] = "=SUM(B7:B12)"
    ws_breakeven['B13'].number_format = currency_format
    
    ws_breakeven['B17'] = "='Доходы и услуги'!D44"
    ws_breakeven['B17'].number_format = currency_format
    
    ws_breakeven['B18'] = "='Доходы и услуги'!F44"
    ws_breakeven['B18'].number_format = currency_format
    
    ws_breakeven['B19'] = "=B18/B17"
    ws_breakeven['B19'].number_format = percent_format
    
    # Форматирование
    for row in range(3, 20):
        for col in range(1, 4):
            cell = ws_breakeven.cell(row=row, column=col)
            if cell.value is not None:
                cell.border = thin_border
    
    for col in range(1, 4):
        cell = ws_breakeven.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    ws_breakeven.column_dimensions['A'].width = 40
    ws_breakeven.column_dimensions['B'].width = 22
    ws_breakeven.column_dimensions['C'].width = 45
    
    # ==================== ЛИСТ 5: Прогноз на 12 месяцев ====================
    ws_forecast = wb.create_sheet(title="Прогноз 12 месяцев")
    
    ws_forecast['A1'] = "Прогноз доходов и расходов на 12 месяцев"
    ws_forecast['A1'].font = title_font
    
    forecast_header = [
        "Показатель", "Месяц 1", "Месяц 2", "Месяц 3", "Месяц 4", 
        "Месяц 5", "Месяц 6", "Месяц 7", "Месяц 8", "Месяц 9", 
        "Месяц 10", "Месяц 11", "Месяц 12", "ИТОГО год"
    ]
    
    forecast_data = [
        ["Выручка всего", 200000, 300000, 400000, 500000, 600000, 650000, 700000, 730000, 760000, 780000, 790000, 800000, None],
        ["— Юр. услуги", 50000, 80000, 120000, 150000, 166000, 170000, 180000, 185000, 190000, 195000, 198000, 200000, None],
        ["— Психология", 50000, 80000, 110000, 140000, 158000, 165000, 170000, 175000, 180000, 185000, 188000, 190000, None],
        ["— Курсы", 60000, 100000, 130000, 170000, 254000, 280000, 300000, 320000, 340000, 350000, 352000, 350000, None],
        ["— Магазин (валовая)", 40000, 40000, 40000, 40000, 94500, 100000, 100000, 100000, 100000, 100000, 100000, 110000, None],
        [],
        ["Расходы всего", 380000, 390000, 400000, 410000, 420000, 425000, 430000, 435000, 440000, 445000, 448000, 450000, None],
        ["Прибыль/убыток", None, None, None, None, None, None, None, None, None, None, None, None, None],
        ["Накопленный результат", None, None, None, None, None, None, None, None, None, None, None, None, None],
    ]
    
    ws_forecast.append([])
    ws_forecast.append(forecast_header)
    for row in forecast_data:
        ws_forecast.append(row)
    
    # Формулы для итогов года
    for row_idx in [1, 2, 3, 4, 5, 7]:
        ws_forecast.cell(row=row_idx+2, column=15).value = f"=SUM(B{row_idx+2}:M{row_idx+2})"
        ws_forecast.cell(row=row_idx+2, column=15).number_format = currency_format
    
    # Формулы для прибыли
    for col_idx in range(2, 15):
        col_letter = get_column_letter(col_idx)
        ws_forecast[f'{col_letter}8'] = f"=B{col_letter}1-{col_letter}7"
        ws_forecast[f'{col_letter}8'].number_format = currency_format
    
    # Формулы для накопленного результата
    ws_forecast['B9'] = "=B8"
    for col_idx in range(3, 15):
        col_letter = get_column_letter(col_idx)
        prev_col = get_column_letter(col_idx - 1)
        ws_forecast[f'{col_letter}9'] = f"={prev_col}9+{col_letter}8"
        ws_forecast[f'{col_letter}9'].number_format = currency_format
    
    # Форматирование
    for row in range(3, 12):
        for col in range(1, 16):
            cell = ws_forecast.cell(row=row, column=col)
            if cell.value is not None:
                cell.border = thin_border
    
    for col in range(1, 16):
        cell = ws_forecast.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Выделение строки прибыли цветом
    for col in range(2, 15):
        col_letter = get_column_letter(col)
        cell = ws_forecast[f'{col_letter}8']
        if col_letter in ['B', 'C']:  # Убыток
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(color="9C0006")
        else:  # Прибыль
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            cell.font = Font(color="006100")
    
    ws_forecast.column_dimensions['A'].width = 25
    for col_idx in range(2, 16):
        col_letter = get_column_letter(col_idx)
        ws_forecast.column_dimensions[col_letter].width = 12
    
    # ==================== ЛИСТ 6: KPI и метрики ====================
    ws_kpi = wb.create_sheet(title="KPI и метрики")
    
    ws_kpi['A1'] = "KPI и метрики успеха"
    ws_kpi['A1'].font = title_font
    
    kpi_header = ["Показатель", "Месяц 1", "Месяц 3", "Месяц 6", "Месяц 12", "Цель", "Статус"]
    kpi_data = [
        ["Выручка, руб.", 200000, 400000, 650000, 800000, 800000, None],
        ["Количество клиентов", 30, 60, 95, 120, 120, None],
        ["Средний чек, руб.", 6700, 6700, 6800, 6700, 7000, None],
        ["Конверсия сайта, %", 1, 2, 2.5, 3, 3.5, None],
        ["LTV (пожизненная ценность), руб.", 15000, 20000, 25000, 30000, 35000, None],
        ["Рентабельность, %", -90, 0, 35, 44, 50, None],
        [],
        ["Источники трафика:"],
        ["— Поисковые системы, %", 30, 35, 40, 45, 50, None],
        ["— Соцсети, %", 40, 35, 30, 25, 20, None],
        ["— Рекомендации, %", 20, 20, 20, 20, 20, None],
        ["— Прямые заходы, %", 10, 10, 10, 10, 10, None],
    ]
    
    ws_kpi.append([])
    ws_kpi.append(kpi_header)
    for row in kpi_data:
        ws_kpi.append(row)
    
    # Формулы для статуса
    for row_idx in range(3, 9):
        goal_col = 'F'
        actual_col = 'E'
        status_col = 'G'
        ws_kpi[f'{status_col}{row_idx}'] = f'=IF({actual_col}{row_idx}>=F{row_idx},"✓","✗")'
        ws_kpi[f'{status_col}{row_idx}'].alignment = Alignment(horizontal='center')
    
    # Форматирование
    for row in range(3, 15):
        for col in range(1, 8):
            cell = ws_kpi.cell(row=row, column=col)
            if cell.value is not None:
                cell.border = thin_border
    
    for col in range(1, 8):
        cell = ws_kpi.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    ws_kpi.column_dimensions['A'].width = 35
    for col_idx in range(2, 8):
        col_letter = get_column_letter(col_idx)
        ws_kpi.column_dimensions[col_letter].width = 12
    
    # ==================== ЛИСТ 7: Источники финансирования ====================
    ws_funding = wb.create_sheet(title="Финансирование")
    
    ws_funding['A1'] = "Источники финансирования"
    ws_funding['A1'].font = title_font
    
    funding_header = ["Источник", "Сумма (руб.)", "Условия", "Срок получения", "Статус"]
    funding_data = [
        ["Личные средства учредителя", 200000, "Безвозмездно", "Немедленно", "Доступно"],
        ["Региональный грант для ветеранов", 300000, "Безвозмездно, софинансирование 25%", "2–3 месяца", "На подаче"],
        ["Соцконтракт (Московская область)", 350000, "Безвозмездно", "1–2 месяца", "На подаче"],
        ["Льготный заём «Мой бизнес»", 500000, "Ставка ~5–7%, отсрочка 6 мес.", "1 месяц", "Резерв"],
        ["ИТОГО доступно", None, "", "", ""],
        [],
        ["Потребность в финансировании:"],
        ["Стартовые вложения", 500000, "Единовременно", "", ""],
        ["Резерв на 2 месяца работы", 800000, "2 × 400 000 руб./мес.", "", ""],
        ["ВСЕГО потребность", None, "", "", ""],
        [],
        ["Профицит/дефицит", None, "", "", ""],
    ]
    
    ws_funding.append([])
    ws_funding.append(funding_header)
    for row in funding_data:
        ws_funding.append(row)
    
    # Формулы
    ws_funding['B6'] = "=SUM(B2:B5)"
    ws_funding['B6'].number_format = currency_format
    
    ws_funding['B10'] = "=B8+B9"
    ws_funding['B10'].number_format = currency_format
    
    ws_funding['B12'] = "=B6-B10"
    ws_funding['B12'].font = Font(bold=True, size=12)
    ws_funding['B12'].number_format = currency_format
    
    # Форматирование
    for row in range(3, 13):
        for col in range(1, 6):
            cell = ws_funding.cell(row=row, column=col)
            if cell.value is not None:
                cell.border = thin_border
    
    for col in range(1, 6):
        cell = ws_funding.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    ws_funding.column_dimensions['A'].width = 45
    ws_funding.column_dimensions['B'].width = 20
    ws_funding.column_dimensions['C'].width = 40
    ws_funding.column_dimensions['D'].width = 18
    ws_funding.column_dimensions['E'].width = 15
    
    # Сохранение файла
    wb.save('финансовая-модель-ООО-ветеранов.xlsx')
    print("✅ Excel-калькулятор создан: финансовая-модель-ООО-ветеранов.xlsx")

if __name__ == '__main__':
    create_financial_model()
